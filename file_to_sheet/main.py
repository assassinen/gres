# import xlrd, xlwt
# rb = xlrd.open_workbook('in/book.xlsx', formatting_info=True)
#
# #выбираем активный лист
# # print(dir(rb))
# # print(rb.sheets)
# # print(rb.user_name)
# # print(rb.sheet_names)
# sheet = rb.sheet_by_index(0)
# # print(dir(sheet))
# val = sheet.row_values(0)
# print(val)

# vals = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]
# print(vals[1][0])

# wb = xlwt.Workbook()
# ws = wb.add_sheet('Test')

#
# #в A1 записываем значение из ячейки A1 прошлого файла
# ws.write(0, 0, vals[1][0])
#
# #в столбец B запишем нашу последовательность из столбца A исходного файла
# # i = 0
# # for rec in vals:
# #     ws.write(i,1, rec[0])
# #     i + i
#
# #сохраняем рабочую книгу
# wb.save('out/xl_rec.xlsx')

import openpyxl
from openpyxl.styles import Font, Alignment

cell_size = {}

def remember_cell_size(cell):
    col_name = cell[0]
    row_name = int(cell[1])
    cell_size[col_name] = source_sheet.column_dimensions[col_name].width
    cell_size[row_name] = source_sheet.row_dimensions[row_name].height

source_wb = openpyxl.load_workbook('in/book.xlsx')
target_wb = openpyxl.Workbook()

source_sheet = source_wb.active
target_sheet = target_wb.active

for row in source_sheet:
    for cell in row:
        remember_cell_size(cell.coordinate)
        font = Font(name=cell.font.name,
                    size=cell.font.size,
                    italic=cell.font.i,
                    bold=cell.font.b,
                    color=cell.font.color,
                    scheme=cell.font.scheme)
        alignment = Alignment(horizontal=cell.alignment.horizontal,
                              vertical=cell.alignment.vertical)
        target_sheet[cell.coordinate].font = font
        target_sheet[cell.coordinate].alignment = alignment
        target_sheet[cell.coordinate] = cell.value

for name, value in cell_size.items():
    if isinstance(name, str):
        target_sheet.column_dimensions[name].width = value
    else:
        target_sheet.row_dimensions[name].height = value


for merge_cell in source_sheet.merged_cells:
    target_sheet.merge_cells(str(merge_cell))


target_wb.save('out/book.xlsx')
# print(cell_size)
