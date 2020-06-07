import openpyxl
from openpyxl.styles import Font, Alignment
import random

cell_size = {}

def remember_cell_size(source_sheet, cell, max_row=0):
    col_index = cell[0]
    row_index = int(cell[1])
    cell_size[col_index] = source_sheet.column_dimensions[col_index].width
    cell_size[int(row_index + max_row)] = source_sheet.row_dimensions[row_index].height
    return f'{col_index}{int(row_index + max_row)}'


def source_to_target(in_path, in_file, target_wb, template_sheets, skip_row=0):
    source_wb = openpyxl.load_workbook(f'{in_path}/{in_file}')
    for sheet in template_sheets:
        try:
            source_sheet = source_wb[sheet]
        except:
            continue
        target_sheet = target_wb[sheet]
        max_row = 0 if in_file == '_template.xlsx' else target_sheet.max_row
        counter = -1

        for row in source_sheet:
            counter += 1
            if counter < skip_row and in_file != '_template.xlsx':
                continue
            for cell in row:
                coordinate = remember_cell_size(source_sheet, cell.coordinate, max_row-skip_row)
                font = Font(name=cell.font.name,
                            size=cell.font.size,
                            italic=cell.font.i,
                            bold=cell.font.b,
                            color=cell.font.color)
                alignment = Alignment(horizontal=cell.alignment.horizontal,
                                      vertical=cell.alignment.vertical,
                                      wrapText=cell.alignment.wrapText)
                target_sheet[coordinate].font = font
                target_sheet[coordinate].alignment = alignment

                target_sheet[coordinate] = cell.value

        for name, value in cell_size.items():
            if isinstance(name, str):
                target_sheet.column_dimensions[name].width = value
            else:
                target_sheet.row_dimensions[name].height = value

        for merge_cell in source_sheet.merged_cells:
            merge = [f'{i[0]}{int(i[1:])+max_row-skip_row}' for i in str(merge_cell).split(':')
                     if int(i[1:])+max_row-skip_row > max_row]
            if len(merge) > 0:
                target_sheet.merge_cells(':'.join(merge))


def add_formulas(target_wb, template_sheets, cells_with_formulas, height_table, skip_row):
    for sheet in template_sheets:
        target_sheet = target_wb[sheet]
        number_items = (target_sheet.max_row - height_table) // (height_table - skip_row) + 1
        for cell in cells_with_formulas:
            cell_value = []
            for number_item in range(1, number_items):
                cell_value += [f'{i[0]}{int(i[1:]) + number_item * (height_table - skip_row)}'
                          for i in str(cell).split(':')]
            if len(cell_value) > 0:
                target_sheet[cell] = '=' + '+'.join(cell_value)


def main():
    in_files = ['tula_book.xlsx', 'orel_book.xlsx', 'klin_book.xlsx']
    in_path = 'in'
    out_file = 'out/result_book.xlsx'

    cells_with_formulas = ['C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5', 'M5', 'N5',
                           'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6',
                           'C7', 'D7', 'E7', 'F7']
    height_table = 7
    skip_row = 4

    template_wb = openpyxl.load_workbook(f'{in_path}/_template.xlsx')
    template_sheets = [worksheets.title for worksheets in template_wb.worksheets]

    target_wb = openpyxl.Workbook()
    for sheet in template_sheets:
        target_wb.create_sheet(title=sheet, index=0)

    source_to_target(in_path, '_template.xlsx', target_wb, template_sheets)

    for in_file in in_files:
        source_to_target(in_path, in_file, target_wb, template_sheets, skip_row)

    add_formulas(target_wb, template_sheets, cells_with_formulas, height_table, skip_row)

    target_wb.remove(target_wb['Sheet'])
    target_wb.save(out_file)


if __name__ == '__main__':
    main()